"""
Reports App - Student 5: Palisha
Handles: Generate PDF/Excel reports — No. of teams, summary, teams without managers
"""
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Count
from teams.models import Team, TeamMember
from organization.models import Department
from django.contrib.auth.models import User
import io
import datetime


@login_required
def reports_home(request):
    """Reports dashboard with summary statistics"""
    total_teams = Team.objects.count()
    active_teams = Team.objects.filter(status='active').count()
    disbanded_teams = Team.objects.filter(status='disbanded').count()
    teams_without_managers = Team.objects.filter(manager__isnull=True).count()
    total_departments = Department.objects.count()
    total_users = User.objects.count()
    dept_stats = Department.objects.annotate(team_count=Count('teams')).order_by('-team_count')

    return render(request, 'reports/reports_home.html', {
        'total_teams': total_teams,
        'active_teams': active_teams,
        'disbanded_teams': disbanded_teams,
        'teams_without_managers': teams_without_managers,
        'total_departments': total_departments,
        'total_users': total_users,
        'dept_stats': dept_stats,
    })


@login_required
def generate_pdf_report(request):
    """Generate a PDF report of all teams"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
    except ImportError:
        return HttpResponse("reportlab not installed. Run: pip install reportlab", status=500)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story = []

    # Title
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, spaceAfter=6)
    story.append(Paragraph("Broadcast Engineering Teams — Report", title_style))
    story.append(Paragraph(f"Generated: {datetime.datetime.now().strftime('%d %B %Y, %H:%M')}", styles['Normal']))
    story.append(Spacer(1, 0.5*cm))

    # Summary section
    story.append(Paragraph("Summary", styles['Heading2']))
    teams = Team.objects.all()
    active = teams.filter(status='active').count()
    no_mgr = teams.filter(manager__isnull=True).count()
    summary_data = [
        ["Metric", "Count"],
        ["Total Teams", str(teams.count())],
        ["Active Teams", str(active)],
        ["Disbanded Teams", str(teams.filter(status='disbanded').count())],
        ["Teams Without Managers", str(no_mgr)],
        ["Total Departments", str(Department.objects.count())],
        ["Total Users", str(User.objects.count())],
    ]
    summary_table = Table(summary_data, colWidths=[10*cm, 5*cm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#042A4A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f4f6f9')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 0.5*cm))

    # Teams table
    story.append(Paragraph("All Teams", styles['Heading2']))
    team_data = [["Team Name", "Department", "Manager", "Members", "Status"]]
    for team in Team.objects.select_related('department', 'manager').prefetch_related('members'):
        team_data.append([
            team.team_name,
            team.department.department_name if team.department else "—",
            team.manager.get_full_name() if team.manager else "No Manager",
            str(team.members.count()),
            team.get_status_display(),
        ])
    team_table = Table(team_data, colWidths=[4.5*cm, 3.5*cm, 4*cm, 2*cm, 3*cm])
    team_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#084070')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f4f6f9')]),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
        ('PADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(team_table)
    story.append(Spacer(1, 0.5*cm))

    # Teams without managers
    story.append(Paragraph("Teams Without Managers", styles['Heading2']))
    orphan_teams = Team.objects.filter(manager__isnull=True)
    if orphan_teams.exists():
        orphan_data = [["Team Name", "Department", "Status"]]
        for t in orphan_teams:
            orphan_data.append([t.team_name, str(t.department or "—"), t.get_status_display()])
        orphan_table = Table(orphan_data, colWidths=[6*cm, 6*cm, 5*cm])
        orphan_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dc3545')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fff5f5')]),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
            ('PADDING', (0, 0), (-1, -1), 5),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
        ]))
        story.append(orphan_table)
    else:
        story.append(Paragraph("All teams have managers assigned.", styles['Normal']))

    doc.build(story)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="broadcast_teams_report.pdf"'
    return response


@login_required
def generate_excel_report(request):
    """Generate an Excel report of all teams"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return HttpResponse("openpyxl not installed. Run: pip install openpyxl", status=500)

    wb = openpyxl.Workbook()

    # ---- Sheet 1: Summary ----
    ws1 = wb.active
    ws1.title = "Summary"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="042A4A")

    ws1['A1'] = "Broadcast Engineering Teams — Summary Report"
    ws1['A1'].font = Font(bold=True, size=14)
    ws1['A2'] = f"Generated: {datetime.datetime.now().strftime('%d %B %Y %H:%M')}"
    ws1.append([])
    ws1.append(["Metric", "Value"])
    for cell in ws1[4]:
        cell.font = header_font
        cell.fill = header_fill

    teams = Team.objects.all()
    rows = [
        ("Total Teams", teams.count()),
        ("Active Teams", teams.filter(status='active').count()),
        ("Disbanded Teams", teams.filter(status='disbanded').count()),
        ("Teams Without Managers", teams.filter(manager__isnull=True).count()),
        ("Total Departments", Department.objects.count()),
        ("Total Users", User.objects.count()),
    ]
    for row in rows:
        ws1.append(row)
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 15

    # ---- Sheet 2: All Teams ----
    ws2 = wb.create_sheet("All Teams")
    headers = ["Team Name", "Department", "Manager", "Contact Email", "Members", "Status", "Purpose"]
    ws2.append(headers)
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor="084070")
    for team in Team.objects.select_related('department', 'manager').prefetch_related('members'):
        ws2.append([
            team.team_name,
            team.department.department_name if team.department else "",
            team.manager.get_full_name() if team.manager else "No Manager",
            team.team_contactemail,
            team.members.count(),
            team.get_status_display(),
            team.team_purpose[:100] if team.team_purpose else "",
        ])
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws2.column_dimensions[col].width = 22

    # ---- Sheet 3: Teams Without Managers ----
    ws3 = wb.create_sheet("No Manager")
    ws3.append(["Team Name", "Department", "Status"])
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor="DC3545")
    for team in Team.objects.filter(manager__isnull=True):
        ws3.append([team.team_name, str(team.department or ""), team.get_status_display()])
    for col in ['A', 'B', 'C']:
        ws3.column_dimensions[col].width = 25

    # ---- Sheet 4: Department Stats ----
    ws4 = wb.create_sheet("Departments")
    ws4.append(["Department", "Specialisation", "Teams Count", "Department Head"])
    for cell in ws4[1]:
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor="042A4A")
    for dept in Department.objects.annotate(tc=Count('teams')):
        ws4.append([
            dept.department_name,
            dept.department_specialisation,
            dept.tc,
            dept.department_head.get_full_name() if dept.department_head else "",
        ])
    for col in ['A', 'B', 'C', 'D']:
        ws4.column_dimensions[col].width = 25

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="broadcast_teams_report.xlsx"'
    return response

def reports_dashboard(request):
    return render(request, 'reports/dashboard.html')
